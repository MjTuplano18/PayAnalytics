"""
Streaming Excel/CSV parser for server-side file uploads.

Uses openpyxl in read-only mode which streams rows without loading the
entire sheet into memory — critical for the 512 MB RAM budget.
"""

import csv
import io
import re
import uuid
from typing import Iterator

from app.schemas.upload import PaymentRecordIn

# Column-name patterns (case-insensitive partial match)
_COL_PATTERNS: dict[str, list[str]] = {
    "bank": ["bank"],
    "account": ["debtor_id", "account", "debtor"],
    "touchpoint": ["tagging", "touchpoint", "tag"],
    "payment_date": ["date_created", "leads_result_edate", "payment date", "edate", "transaction date", "trans_date", "value_date", "posting_date", "date_paid"],
    "payment_amount": ["leads_result_amount", "payment amount", "amount"],
    "environment": ["environment", "env"],
    "month": ["month"],
}


def _match_column(header: str, patterns: list[str]) -> bool:
    lower = header.lower()
    return any(p in lower for p in patterns)


def _resolve_columns(headers: list[str]) -> dict[str, int | None]:
    """Map logical field names to column indices."""
    mapping: dict[str, int | None] = {}
    for field, patterns in _COL_PATTERNS.items():
        idx = next(
            (i for i, h in enumerate(headers) if _match_column(h, patterns)),
            None,
        )
        mapping[field] = idx
    return mapping


def _format_date(value: object) -> str:
    """Convert cell values to YYYY-MM-DD string."""
    if value is None:
        return ""
    from datetime import datetime, date

    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    
    s = str(value).strip()
    
    # Try to parse common date formats
    # Format: MM/DD/YYYY or M/D/YYYY (e.g., "01/12/2026" or "1/12/2026")
    if "/" in s:
        try:
            parts = s.split("/")
            if len(parts) == 3:
                month, day, year = parts
                # Handle 2-digit or 4-digit year
                if len(year) == 2:
                    year = "20" + year
                dt = datetime(int(year), int(month), int(day))
                return dt.strftime("%Y-%m-%d")
        except (ValueError, IndexError):
            pass
    
    # Format: DD-MM-YYYY or D-M-YYYY
    if "-" in s and not re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        try:
            parts = s.split("-")
            if len(parts) == 3:
                day, month, year = parts
                if len(year) == 2:
                    year = "20" + year
                dt = datetime(int(year), int(month), int(day))
                return dt.strftime("%Y-%m-%d")
        except (ValueError, IndexError):
            pass
    
    # Excel serial number (integer between 30000 and 100000)
    if re.match(r"^\d+$", s):
        num = int(s)
        if 30000 < num < 100000:
            from datetime import timedelta

            epoch = datetime(1899, 12, 30)
            dt = epoch + timedelta(days=num)
            return dt.strftime("%Y-%m-%d")
    
    # Already in YYYY-MM-DD format
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    
    # If all parsing fails, return the original string
    return s


def _safe_float(value: object) -> float:
    from decimal import Decimal, InvalidOperation
    try:
        return float(Decimal(str(value)).quantize(Decimal("0.01")))
    except (TypeError, ValueError, InvalidOperation):
        return 0.0


def _row_to_record(
    values: list[object], col_map: dict[str, int | None]
) -> PaymentRecordIn | None:
    def _get(field: str) -> object:
        idx = col_map.get(field)
        if idx is None or idx >= len(values):
            return None
        return values[idx]

    bank_raw = _get("bank")
    account_raw = _get("account")

    # Skip rows where both bank and account are empty (e.g. total/summary rows)
    if not bank_raw and not account_raw:
        return None

    return PaymentRecordIn(
        bank=str(bank_raw or "Unknown"),
        account=str(account_raw or "") or f"NO-ACCT-{uuid.uuid4().hex[:8].upper()}",
        touchpoint=str(_get("touchpoint") or "NO TOUCHPOINT"),
        payment_date=_format_date(_get("payment_date")),
        payment_amount=_safe_float(_get("payment_amount")),
        environment=str(_get("environment") or "") if _get("environment") else None,
        month=str(_get("month") or "") if _get("month") else None,
    )


def stream_xlsx(file_bytes: bytes) -> Iterator[PaymentRecordIn]:
    """
    Parse an .xlsx file in streaming (read-only) mode.
    Yields one PaymentRecordIn per data row.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    col_map: dict[str, int | None] | None = None
    for row in ws.iter_rows(values_only=True):
        if col_map is None:
            # First row = headers
            headers = [str(c or "") for c in row]
            col_map = _resolve_columns(headers)
            continue
        # Skip completely empty rows
        if all(c is None for c in row):
            continue
        record = _row_to_record(list(row), col_map)
        if record is not None:
            yield record

    wb.close()


def stream_csv(file_bytes: bytes) -> Iterator[PaymentRecordIn]:
    """Parse a CSV file and yield PaymentRecordIn per row."""
    text = file_bytes.decode("utf-8-sig")  # handle BOM
    reader = csv.reader(io.StringIO(text))
    col_map: dict[str, int | None] | None = None
    for row_values in reader:
        if col_map is None:
            col_map = _resolve_columns(row_values)
            continue
        if not any(v.strip() for v in row_values):
            continue
        record = _row_to_record(list(row_values), col_map)
        if record is not None:
            yield record
